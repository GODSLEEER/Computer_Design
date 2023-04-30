import pydot
import numpy as np
import pandas as pd
import scipy.stats as stats
import matplotlib.pyplot as plt
from sklearn import metrics
from openpyxl import load_workbook
from sklearn.tree import export_graphviz
from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import train_test_split
import joblib


# -------------------------------------------------Data Process----------------------------------------------------
random_forest_seed=60
random_seed=44
write_excel_path='Result List/ParameterResult_ML.xlsx'
tree_graph_dot_path='Result List/tree.dot'
tree_graph_png_path='Result List/tree.png'

Train_Feature_Name=['Length (ft)','YearDis','IsCatamarans','LOA','LWL','Beam','SA','Draft','Displacement','GDP','Gini']
df_raw = pd.read_csv('Final.csv')
X1 = df_raw.loc[:, ['Country/Region/State ']]
X1=pd.get_dummies(X1,columns=["Country/Region/State "])   #get_dummies对“整数特征”无变化，对“类别特征”one-hot编码
Make_Feature=list(X1)
Rigion_Name=[x[22:] for x in Make_Feature]
X1=X1.values
X2=df_raw.loc[:,Train_Feature_Name ].values
X = np.hstack([X2,X1])
Train_Feature_Name=Train_Feature_Name+Make_Feature
y= df_raw.loc[:, 'Listing Price (USD)'].values
x_train, x_test, y_train, y_test = train_test_split(X, y, random_state=random_forest_seed, train_size=0.9)

# -----------------------------------------------------Model Train-------------------------------------------------
forest1 = RandomForestRegressor(n_estimators=200,max_depth=38)
forest1.fit(x_train, y_train)
print("Traing Score:%f" % forest1.score(x_train, y_train))
print("Testing Score:%f" % forest1.score(x_test, y_test))
joblib.dump(forest1, 'Regreesion_Price_Predict.model')
forest_predict=forest1.predict(x_test)
forest_error=forest_predict-y_test
# ----------------------------------------------------Draw test plot--------------------------------------------------

# plt.figure(1)
# plt.clf()
# ax=plt.axes(aspect='equal')
# plt.xlabel('True Values')
# plt.ylabel('Predictions')
# Lims=[0,1000000]
# plt.xlim(Lims)
# plt.ylim(Lims)
# plt.plot(Lims,Lims)
# plt.scatter(y_test,forest_predict)
# plt.grid(False)
# plt.show()

# plt.figure(2)
# plt.clf()
# plt.hist(forest_error,bins=30)
# plt.xlabel('Prediction Error')
# plt.ylabel('Count')
# plt.grid(False)
# plt.show()


# ----------------------------------------------Verify the accuracy---------------------------------------

random_forest_pearson_r=stats.pearsonr(y_test,forest_predict)
random_forest_R2=metrics.r2_score(y_test,forest_predict)
random_forest_RMSE=metrics.mean_squared_error(y_test,forest_predict)**0.5
print('Pearson correlation coefficient is {0}, and RMSE is {1}.'.format(random_forest_pearson_r[0],
                                                                        random_forest_RMSE))

# -----------------------------------------------Save key parameters-----------------------------------------


excel_file=load_workbook(write_excel_path)
excel_all_sheet=excel_file.sheetnames
excel_write_sheet=excel_file[excel_all_sheet[0]]
excel_write_sheet=excel_file.active
max_row=excel_write_sheet.max_row
excel_write_content=[random_forest_pearson_r[0],random_forest_R2,random_forest_RMSE,random_seed,random_forest_seed]
for i in range(len(excel_write_content)):
        exec("excel_write_sheet.cell(max_row+1,i+1).value=excel_write_content[i]")
excel_file.save(write_excel_path)

# -----------------------------------------------Draw decision tree visualizing plot---------------------------------

random_forest_tree=forest1.estimators_[5]
export_graphviz(random_forest_tree,out_file=tree_graph_dot_path,
                feature_names=Train_Feature_Name,rounded=True,precision=1)
(random_forest_graph,)=pydot.graph_from_dot_file(tree_graph_dot_path)
random_forest_graph.write_png(tree_graph_png_path)

# ---------------------------------------------Calculate the importance of variables---------------------------------

random_forest_importance=list(forest1.feature_importances_)[11:]
random_forest_feature_importance=[(feature,round(importance,8)) 
                                  for feature, importance in zip(Rigion_Name,random_forest_importance)]
random_forest_feature_importance=sorted(random_forest_feature_importance,key=lambda x:x[1],reverse=True)
plt.figure(3)
plt.figure(i).set_size_inches(18.5, 10.5)
plt.clf()
importance_plot_x_values=list(range(len(random_forest_importance)))
plt.bar(importance_plot_x_values,random_forest_importance,orientation='vertical')
plt.xticks(importance_plot_x_values,Rigion_Name,rotation='vertical')
plt.xlabel('Variable')
plt.ylabel('Importance')
plt.title('Variable Importances')
save_path='Result List/ImportanceVariable'+str(i)+'.png'
plt.savefig(save_path,bbox_inches='tight',dpi=100)
plt.show()