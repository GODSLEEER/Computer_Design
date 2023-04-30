import numpy as np
import openpyxl
import joblib
from sklearn import metrics
import matplotlib.pyplot as plt
import scipy.stats as stats
import pandas as pd

def split_list_n_list(origin_list, n):
    if len(origin_list) % n == 0:
        cnt = len(origin_list) // n
    else:
        cnt = len(origin_list) // n + 1
    lists=[]
    for i in range(0, n):
        lists.append(origin_list[i*cnt:(i+1)*cnt])
    return lists

file_path0='Hong Kong.xlsx'
file_path1='Result List/Hong Kong Prediction_List.xlsx'
Rpp = joblib.load('Regreesion_Price_Predict.model')


Make_Var0_Pos_xlsx = openpyxl.load_workbook(file_path0)
Make_Var0_sheet = Make_Var0_Pos_xlsx['Sheet1']   
Make_Var0_row_num=Make_Var0_sheet.max_row
Make_Var0_col_num=Make_Var0_sheet.max_column

Region_sheet = Make_Var0_Pos_xlsx['Sheet2']   
Region_row_num=Region_sheet.max_row
Region_col_num=Region_sheet.max_column

Region_List=[]
Eco_List=[]
for i in range(2, Region_row_num+1):
    Region_List.append(Region_sheet.cell(i,2).value)
    Eco_List.append([Region_sheet.cell(i,3).value,Region_sheet.cell(i,4).value])


df_raw = pd.read_csv('Final.csv')
X1 = df_raw.loc[:, ['Make']]
X1=pd.get_dummies(X1,columns=["Make"])   #get_dummies对“整数特征”无变化，对“类别特征”one-hot编码
Make_Feature=list(X1)
Make_to_Encode={}
for i in range(len(Make_Feature)):
    Make_to_Encode[Make_Feature[i]]=i


Name_List=[]
Double_Predict_List=[[],[]]   #0单体，1双体
Double_Gini_List=[[],[]]
Double_GDP_List=[[],[]]
Is_Hongkong_List=[[],[]]
for i in range(2,Make_Var0_row_num+1):
    Make = Make_Var0_sheet.cell(i,1).value
    Variant= Make_Var0_sheet.cell(i,2).value
    Name= str(Make) + " " + str(Variant)
    Length= Make_Var0_sheet.cell(i,3).value
    Region= Make_Var0_sheet.cell(i,5).value
    IsCatamarans= int(Make_Var0_sheet.cell(i,8).value)
    LOA= Make_Var0_sheet.cell(i,9).value
    LWL= Make_Var0_sheet.cell(i,10).value
    Beam= Make_Var0_sheet.cell(i,11).value
    SA= Make_Var0_sheet.cell(i,12).value
    Draft= Make_Var0_sheet.cell(i,13).value
    Displacement= Make_Var0_sheet.cell(i,14).value
    GDP= Make_Var0_sheet.cell(i,15).value
    Gini= Make_Var0_sheet.cell(i,16).value
    YearDis=Make_Var0_sheet.cell(i,17).value
    Price= Make_Var0_sheet.cell(i,6).value
    Encode_List=[0 for i in range(73)]
    Encode_List[Make_to_Encode['Make'+"_"+str(Make)]]=1
    List_Data=[Length,YearDis,IsCatamarans,LOA,LWL,Beam,SA,Draft,Displacement,GDP,Gini]
    Input_List=List_Data+Encode_List
    Predict_Price=Rpp.predict(np.array(Input_List).reshape((- 1, 84))).item()
    Double_Predict_List[IsCatamarans].append(Price)
    Double_Gini_List[IsCatamarans].append(Gini)
    Double_GDP_List[IsCatamarans].append(GDP)
    Is_Hongkong_List[IsCatamarans].append(1)
    for i in range(len(Region_List)):
        List_Data[-2]=Eco_List[i][0]
        List_Data[-1]=Eco_List[i][1]
        Input_List=List_Data+Encode_List
        Predict_Price=Rpp.predict(np.array(Input_List).reshape((- 1, 84))).item()
        Double_Predict_List[IsCatamarans].append(Predict_Price)
        Double_Gini_List[IsCatamarans].append(Gini)
        Double_GDP_List[IsCatamarans].append(GDP)
        Is_Hongkong_List[IsCatamarans].append(0)

        
        


Predict_List=Double_Predict_List[0]+Double_Predict_List[1]

spearmanr_Hongkong_Single=stats.spearmanr(Is_Hongkong_List[0],Double_Predict_List[0])[0]
spearmanr_Hongkong_Double=stats.spearmanr(Is_Hongkong_List[1],Double_Predict_List[1])[0]

Spearmanr_Hongkong_List=[spearmanr_Hongkong_Single,spearmanr_Hongkong_Double]

plot_x_values=list(range(2))

plt.figure(1)
plt.clf()
plt.bar(plot_x_values,Spearmanr_Hongkong_List,orientation='vertical')
plt.xticks(plot_x_values,['Single', 'Double'],rotation='vertical')
plt.xlabel('IsCatamarans')
plt.ylabel('spearmanr_Hongkong')
plt.title('spearmanr_Hongkong')
save_path='figs/spearmanr_Hongkong'+str(1)+'.png'
plt.savefig(save_path,bbox_inches='tight',dpi=100)
plt.show()
