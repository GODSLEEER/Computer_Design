import numpy as np
import openpyxl
import joblib
from sklearn import metrics
import pandas as pd

varid=int(input("请输入您要预测的二手船种类id："))
YearDis=2020-int(input("请输入船只的制造年份："))
Region=input("请输入船只的产地：")

file_path0='Final.xlsx'
file_path2='make and variant.xlsx'
file_path3='place and economic.xlsx'
Rpp = joblib.load('Regreesion_Price_Predict.model')


Make_Var0_Pos_xlsx = openpyxl.load_workbook(file_path0)
Make_Var0_sheet = Make_Var0_Pos_xlsx['Sheet1']
Make_Var0_row_num=Make_Var0_sheet.max_row
Make_Var0_col_num=Make_Var0_sheet.max_column

given_xlsx = openpyxl.load_workbook(file_path2)
given_sheet = given_xlsx['Sheet1']
given_row_num=given_sheet.max_row
given_col_num=given_sheet.max_column

eco_xlsx = openpyxl.load_workbook(file_path3)
eco_sheet = eco_xlsx['Sheet1']
eco_row_num=eco_sheet.max_row
eco_col_num=eco_sheet.max_column

Make=given_sheet.cell(varid+1,1).value
Variant=given_sheet.cell(varid+1,2).value

#建立Make的one-hot对应表
df_raw = pd.read_csv('Final.csv')
X1 = df_raw.loc[:, ['Make']]
X1=pd.get_dummies(X1,columns=["Make"])   #get_dummies对“整数特征”无变化，对“类别特征”one-hot编码
Make_Feature=list(X1)
Make_to_Encode={}
for i in range(len(Make_Feature)):
    Make_to_Encode[Make_Feature[i]]=i
Encode_List=[0 for i in range(73)]
Encode_List[Make_to_Encode['Make'+"_"+str(Make)]]=1

for i in range(2,Make_Var0_row_num):
    Make_tmp = Make_Var0_sheet.cell(i,1).value
    Variant_tmp= Make_Var0_sheet.cell(i,2).value
    if(Make_tmp==Make and Variant_tmp==Variant):
        IsCatamarans= Make_Var0_sheet.cell(i,8).value
        Length= Make_Var0_sheet.cell(i,3).value
        LOA= Make_Var0_sheet.cell(i,9).value
        LWL= Make_Var0_sheet.cell(i,10).value
        Beam= Make_Var0_sheet.cell(i,11).value
        SA= Make_Var0_sheet.cell(i,12).value
        Draft= Make_Var0_sheet.cell(i,13).value
        Displacement= Make_Var0_sheet.cell(i,14).value
        break

for i in range(2,eco_row_num+1):
    Region_tmp=eco_sheet.cell(i,2).value
    if(Region.upper()==Region_tmp.upper()):
        GDP=eco_sheet.cell(i,3).value
        Gini=eco_sheet.cell(i,4).value


List_Data=[Length,YearDis,IsCatamarans,LOA,LWL,Beam,SA,Draft,Displacement,GDP,Gini]
Input_List=List_Data+Encode_List
Predict_Price=Rpp.predict(np.array(Input_List).reshape((- 1, 84))).item()
print("本模型预测该二手船价格为：",Predict_Price,"Dollar")