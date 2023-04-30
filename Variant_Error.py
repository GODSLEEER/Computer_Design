import numpy as np
import openpyxl
import joblib
from sklearn import metrics
import matplotlib.pyplot as plt
import pandas as pd

def split_list_n_list(origin_list, n):
    if len(origin_list) % n == 0:
        cnt = len(origin_list) // n
    else:
        cnt = len(origin_list) // n + 1
    lists=[]
    for i in range(0, n):
        lists.append(origin_list[i*cnt:(i+1)*cnt])
    return lists

file_path0='Final.xlsx'
file_path1='Variant_AvePrice.xlsx'
Rpp = joblib.load('Regreesion_Price_Predict.model')

Make_Var0_Pos_xlsx = openpyxl.load_workbook(file_path0)
Make_Var0_sheet = Make_Var0_Pos_xlsx['Sheet1']   
Make_Var0_row_num=Make_Var0_sheet.max_row
Make_Var0_col_num=Make_Var0_sheet.max_column

given_xlsx = openpyxl.load_workbook(file_path1)
given_sheet = given_xlsx['Sheet1']   

df_raw = pd.read_csv('Final.csv')
X1 = df_raw.loc[:, ['Make']]
X1=pd.get_dummies(X1,columns=["Make"])   #get_dummies对“整数特征”无变化，对“类别特征”one-hot编码
Make_Feature=list(X1)
Make_to_Encode={}
for i in range(len(Make_Feature)):
    Make_to_Encode[Make_Feature[i]]=i

Variant_Price_Dict={}
Variant_Prediction_Dict={}
Name_List=[]
for i in range(2,Make_Var0_row_num):
    Make = Make_Var0_sheet.cell(i,1).value
    Variant= Make_Var0_sheet.cell(i,2).value
    Name= str(Make) + " " + str(Variant)
    Length= Make_Var0_sheet.cell(i,3).value
    Region= Make_Var0_sheet.cell(i,5).value
    IsCatamarans= Make_Var0_sheet.cell(i,8).value
    LOA= Make_Var0_sheet.cell(i,9).value
    LWL= Make_Var0_sheet.cell(i,10).value
    Beam= Make_Var0_sheet.cell(i,11).value
    SA= Make_Var0_sheet.cell(i,12).value
    Draft= Make_Var0_sheet.cell(i,13).value
    Displacement= Make_Var0_sheet.cell(i,14).value
    GDP= Make_Var0_sheet.cell(i,15).value
    Gini= Make_Var0_sheet.cell(i,16).value
    YearDis=Make_Var0_sheet.cell(i,17).value
    Price= Make_Var0_sheet.cell(i,6).value
    List_Data=[Length,YearDis,IsCatamarans,LOA,LWL,Beam,SA,Draft,Displacement,GDP,Gini]
    Encode_List=[0 for i in range(73)]
    Encode_List[Make_to_Encode['Make'+"_"+str(Make)]]=1
    Input_List=List_Data+Encode_List
    Predict_Price=Rpp.predict(np.array(Input_List).reshape((- 1, 84))).item()
    if Name not in Variant_Price_Dict:
        Name_List.append(Name)
        Variant_Price_Dict[Name]=[Price]
        Variant_Prediction_Dict[Name]=[Predict_Price]
    else:
        Variant_Price_Dict[Name].append(Price)
        Variant_Prediction_Dict[Name].append(Predict_Price)

Percentage_List=[]
for nm in range(len(Name_List)):
    length=len(Variant_Price_Dict[Name_List[nm]])
    errors= [Variant_Price_Dict[Name_List[nm]][i]-Variant_Prediction_Dict[Name_List[nm]][i] for i in range(length)]  
    error=0
    Price_Sum=0
    for i in range(0,length):
        error+=abs(errors[i])
        Price_Sum+=Variant_Price_Dict[Name_List[nm]][i]
    percentage_error=error/Price_Sum
    Percentage_List.append(percentage_error)
    given_sheet.cell(nm+2,3).value=percentage_error

given_xlsx.save(file_path1)

plot_x_values=list(range(len(Percentage_List)))

R_Square_List_Sp=split_list_n_list(Percentage_List,7)
plot_x_values_Sp=split_list_n_list(plot_x_values,7)
Name_List_Sp = split_list_n_list(Name_List,7)

for i in range(7):
    plt.figure(i).set_size_inches(18.5, 10.5)
    plt.clf()
    plt.bar(plot_x_values_Sp[i],R_Square_List_Sp[i],orientation='vertical')
    plt.xticks(plot_x_values_Sp[i],Name_List_Sp[i],rotation='vertical')
    plt.xlabel('Variant Name')
    plt.ylabel('Percentage Error ')
    plt.title('Percentage Error Each Variant')
    save_path='figs/savefig_example'+str(i)+'.png'
    plt.savefig(save_path,bbox_inches='tight',dpi=100)
    plt.show()


